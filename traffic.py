from google.colab import files

f = files.upload( )

import openpyxl

wb = openpyxl.Workbook( )
ws = wb.active

import csv

with open('data1.csv', encoding = 'cp949') as f:
  rdr = csv.reader(f)
  for line in rdr:
    ws.append(line)

ws2 = wb.create_sheet('Sheet2')

with open('data2.csv', encoding = 'cp949') as f:
  rdr = csv.reader(f)
  for line in rdr:
    ws2.append(line)

ws3 = wb.create_sheet('Sheet3')

with open('data3.csv', encoding = 'cp949') as f:
  rdr = csv.reader(f)
  for line in rdr:
    ws3.append(line)

ws.title = '경기도_의정부시_의정부경전철 혼잡도'
ws2.title = '대전교통공사_열차 혼잡도 분석'
ws3.title = '서울특별시_지하철 혼잡도 정보'

wb.save('traffic.xlsx')

